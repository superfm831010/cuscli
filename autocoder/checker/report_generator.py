"""
代码检查报告生成器

支持生成 JSON、Markdown 和 Excel (xlsx) 格式的检查报告。
"""

import json
import os
import hashlib
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from loguru import logger

# Excel 导出支持
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from autocoder.checker.types import (
    FileCheckResult,
    BatchCheckResult,
    GitInfo,
    Issue,
    Severity
)


class ReportGenerator:
    """
    报告生成器

    支持生成单文件报告和批量检查汇总报告，支持 JSON、Markdown 和 Excel 三种格式。
    默认使用 Excel 格式，可通过 set_export_format() 方法切换。

    报告目录结构:
        codecheck/
        └── {check_id}_YYYYMMDD_HHMMSS/
            ├── check.log             # 检查任务日志（详细执行过程）
            ├── summary.xlsx          # 批量检查汇总（Excel，默认）
            └── files/
                ├── with_issues/      # 有问题的文件报告
                │   ├── file1_py.xlsx
                │   └── ...
                └── no_issues/        # 无问题的文件报告
                    ├── file2_py.xlsx
                    └── ...

    Attributes:
        output_dir: 报告输出根目录
        export_format: 导出格式 ("xlsx"、"json" 或 "md")
    """

    # ==================== Excel 样式定义 ====================

    # 严重程度颜色（浅色背景，便于阅读）
    SEVERITY_FILLS = {
        "error": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),    # 浅红色
        "warning": PatternFill(start_color="FFEBCC", end_color="FFEBCC", fill_type="solid"),  # 浅橙色
        "info": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),     # 浅蓝色
    }

    # 表头样式（深蓝背景 + 白色加粗字体）
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    # 检查状态颜色
    STATUS_FILLS = {
        "success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 浅绿色
        "failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),   # 浅红色
        "skipped": PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid"),  # 浅灰色
        "timeout": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 浅黄色
    }

    # 细边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, output_dir: str = "codecheck"):
        """
        初始化报告生成器

        Args:
            output_dir: 报告输出根目录，默认为 "codecheck"
        """
        self.output_dir = output_dir
        # 默认导出格式为 Excel
        self.export_format = "xlsx"
        logger.info(f"报告生成器已初始化，输出目录: {output_dir}, 格式: {self.export_format}")

    def set_export_format(self, format_type: str) -> None:
        """
        设置导出格式

        Args:
            format_type: 导出格式，可选 "xlsx"（默认）、"json" 或 "md"

        Raises:
            ValueError: 如果格式不支持
        """
        format_type = format_type.lower()
        if format_type not in ("xlsx", "json", "md"):
            raise ValueError(f"不支持的导出格式: {format_type}，可选: xlsx, json, md")
        self.export_format = format_type
        logger.info(f"报告导出格式已设置为: {format_type}")

    def generate_file_report(
        self, result: FileCheckResult, report_dir: str
    ) -> None:
        """
        生成单个文件的检查报告

        根据 export_format 设置选择输出格式：
        - "xlsx": Excel 格式（默认）
        - "json": JSON 格式
        - "md": Markdown 格式

        Args:
            result: 文件检查结果
            report_dir: 报告输出目录

        Raises:
            RuntimeError: 如果报告生成失败
        """
        try:
            # 根据是否有问题决定保存到哪个子目录
            has_issues = result.get_total_issues() > 0
            subdir = "with_issues" if has_issues else "no_issues"

            # 创建对应的子目录
            files_dir = os.path.join(report_dir, "files", subdir)
            os.makedirs(files_dir, exist_ok=True)

            # 生成安全的文件名
            safe_filename = self._safe_path(result.file_path)

            # 根据导出格式生成对应的报告
            if self.export_format == "xlsx":
                # Excel 格式（默认）
                xlsx_path = os.path.join(files_dir, f"{safe_filename}.xlsx")
                self._generate_excel_file_report(result, xlsx_path)
                logger.info(f"已生成 Excel 文件报告: {result.file_path} -> {subdir}")

            elif self.export_format == "json":
                # JSON 格式
                json_path = os.path.join(files_dir, f"{safe_filename}.json")
                self._generate_json_report(result, json_path)

                # 验证 JSON 文件是否真的创建成功
                if not os.path.exists(json_path):
                    raise RuntimeError(
                        f"JSON 报告文件未创建成功: {json_path}\n"
                        f"可能原因：权限不足、磁盘空间不足或路径包含特殊字符"
                    )
                logger.info(f"已生成 JSON 文件报告: {result.file_path} -> {subdir}")

            elif self.export_format == "md":
                # Markdown 格式
                md_path = os.path.join(files_dir, f"{safe_filename}.md")
                md_content = self._format_file_markdown(result)
                self._generate_markdown_report(md_content, md_path)

                # 验证 Markdown 文件是否真的创建成功
                if not os.path.exists(md_path):
                    raise RuntimeError(
                        f"Markdown 报告文件未创建成功: {md_path}\n"
                        f"可能原因：权限不足、磁盘空间不足或路径包含特殊字符"
                    )
                logger.info(f"已生成 Markdown 文件报告: {result.file_path} -> {subdir}")

        except Exception as e:
            error_msg = f"生成文件报告失败 {result.file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            # 重新抛出异常，让调用者知道报告生成失败
            raise RuntimeError(error_msg) from e

    def generate_summary_report(
        self, results: List[FileCheckResult], report_dir: str, git_info: Optional[GitInfo] = None
    ) -> None:
        """
        生成批量检查的汇总报告

        根据 export_format 设置选择输出格式：
        - "xlsx": Excel 格式（默认）
        - "json": JSON 格式
        - "md": Markdown 格式

        Args:
            results: 所有文件的检查结果列表
            report_dir: 报告输出目录
            git_info: Git 检查信息（可选，Phase 4 新增）
        """
        try:
            # 确保报告目录存在
            os.makedirs(report_dir, exist_ok=True)

            # 根据导出格式生成对应的报告
            if self.export_format == "xlsx":
                # Excel 格式（默认）
                xlsx_path = os.path.join(report_dir, "summary.xlsx")
                self._generate_excel_summary_report(results, xlsx_path, git_info)
                logger.info(f"已生成 Excel 汇总报告: {report_dir}")

            elif self.export_format == "json":
                # JSON 格式 - 需要构建 BatchCheckResult
                start_time = datetime.now().isoformat()
                end_time = datetime.now().isoformat()

                total_issues = sum(len(r.issues) for r in results)
                total_errors = sum(r.error_count for r in results)
                total_warnings = sum(r.warning_count for r in results)
                total_infos = sum(r.info_count for r in results)

                batch_result = BatchCheckResult(
                    check_id=os.path.basename(report_dir),
                    start_time=start_time,
                    end_time=end_time,
                    total_files=len(results),
                    checked_files=len([r for r in results if r.status == "success"]),
                    total_issues=total_issues,
                    total_errors=total_errors,
                    total_warnings=total_warnings,
                    total_infos=total_infos,
                    file_results=results,
                    git_info=git_info
                )

                json_path = os.path.join(report_dir, "summary.json")
                self._generate_json_report(batch_result, json_path)
                logger.info(f"已生成 JSON 汇总报告: {report_dir}")

            elif self.export_format == "md":
                # Markdown 格式 - 需要构建 BatchCheckResult
                start_time = datetime.now().isoformat()
                end_time = datetime.now().isoformat()

                total_issues = sum(len(r.issues) for r in results)
                total_errors = sum(r.error_count for r in results)
                total_warnings = sum(r.warning_count for r in results)
                total_infos = sum(r.info_count for r in results)

                batch_result = BatchCheckResult(
                    check_id=os.path.basename(report_dir),
                    start_time=start_time,
                    end_time=end_time,
                    total_files=len(results),
                    checked_files=len([r for r in results if r.status == "success"]),
                    total_issues=total_issues,
                    total_errors=total_errors,
                    total_warnings=total_warnings,
                    total_infos=total_infos,
                    file_results=results,
                    git_info=git_info
                )

                md_path = os.path.join(report_dir, "summary.md")
                md_content = self._format_summary_markdown(batch_result)
                self._generate_markdown_report(md_content, md_path)
                logger.info(f"已生成 Markdown 汇总报告: {report_dir}")

        except Exception as e:
            logger.error(f"生成汇总报告失败: {e}", exc_info=True)

    def _generate_json_report(self, data: Any, output_path: str) -> None:
        """
        生成 JSON 格式报告

        Args:
            data: 要保存的数据（支持 pydantic 模型）
            output_path: 输出文件路径

        Raises:
            OSError: 如果文件写入失败
            RuntimeError: 如果文件创建后验证失败
        """
        try:
            # 确保目录存在
            parent_dir = os.path.dirname(output_path)
            if parent_dir:  # 避免空字符串导致的问题
                os.makedirs(parent_dir, exist_ok=True)

            # 如果是 pydantic 模型，使用 model_dump
            if hasattr(data, 'model_dump'):
                json_data = data.model_dump()
            else:
                json_data = data

            # 写入 JSON 文件
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=4)

            # 验证文件是否真的创建成功
            if not os.path.exists(output_path):
                raise RuntimeError(
                    f"JSON 文件写入后验证失败，文件不存在: {output_path}\n"
                    f"可能原因：写入时发生错误但未抛出异常"
                )

            # 验证文件大小
            file_size = os.path.getsize(output_path)
            if file_size == 0:
                raise RuntimeError(
                    f"JSON 文件写入后验证失败，文件大小为 0: {output_path}"
                )

            logger.debug(f"JSON 报告已保存: {output_path} (大小: {file_size} 字节)")

        except (OSError, IOError) as e:
            error_msg = f"生成 JSON 报告失败 {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise OSError(error_msg) from e
        except Exception as e:
            error_msg = f"生成 JSON 报告时发生未预期的错误 {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise

    def _generate_markdown_report(self, content: str, output_path: str) -> None:
        """
        生成 Markdown 格式报告

        Args:
            content: Markdown 内容
            output_path: 输出文件路径

        Raises:
            OSError: 如果文件写入失败
            RuntimeError: 如果文件创建后验证失败
        """
        try:
            # 确保目录存在
            parent_dir = os.path.dirname(output_path)
            if parent_dir:  # 避免空字符串导致的问题
                os.makedirs(parent_dir, exist_ok=True)

            # 写入 Markdown 文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(content)

            # 验证文件是否真的创建成功
            if not os.path.exists(output_path):
                raise RuntimeError(
                    f"Markdown 文件写入后验证失败，文件不存在: {output_path}\n"
                    f"可能原因：写入时发生错误但未抛出异常"
                )

            # 验证文件大小
            file_size = os.path.getsize(output_path)
            if file_size == 0:
                raise RuntimeError(
                    f"Markdown 文件写入后验证失败，文件大小为 0: {output_path}"
                )

            logger.debug(f"Markdown 报告已保存: {output_path} (大小: {file_size} 字节)")

        except (OSError, IOError) as e:
            error_msg = f"生成 Markdown 报告失败 {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise OSError(error_msg) from e
        except Exception as e:
            error_msg = f"生成 Markdown 报告时发生未预期的错误 {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise

    def _format_file_markdown(self, result: FileCheckResult) -> str:
        """
        格式化单文件检查结果为 Markdown

        Args:
            result: 文件检查结果

        Returns:
            Markdown 格式的报告内容
        """
        # 确定状态图标
        status_icon = {
            "success": "✅",
            "failed": "❌",
            "skipped": "⏭️"
        }.get(result.status, "❓")

        # 构建 Markdown 内容
        md = f"""# 📄 文件检查报告

**文件路径**: `{result.file_path}`
**检查时间**: {result.check_time}
**检查状态**: {status_icon} {result.status}
"""

        # Phase 5: 显示审核模式和统计信息
        if result.audit_mode:
            audit_mode_icon = "🎯" if result.audit_mode == "diff-only" else "📄"
            audit_mode_text = "Diff-Only（差异审核）" if result.audit_mode == "diff-only" else "全文件审核"
            md += f"**审核模式**: {audit_mode_icon} {audit_mode_text}\n"

            if result.audit_stats and result.audit_mode == "diff-only":
                audit_summary = result.get_audit_summary()
                if audit_summary:
                    md += f"**审核范围**: {audit_summary}\n"

        md += f"**问题总数**: {result.get_total_issues()} 个\n\n"
        md += f"""## 📊 问题统计

| 严重程度 | 数量 |
|---------|------|
| ❌ 错误 (ERROR) | {result.error_count} |
| ⚠️ 警告 (WARNING) | {result.warning_count} |
| ℹ️ 提示 (INFO) | {result.info_count} |
| **总计** | **{result.get_total_issues()}** |

---

"""

        # 如果检查失败，显示错误信息
        if result.status == "failed" and result.error_message:
            md += f"""## ❌ 检查错误

```
{result.error_message}
```

---

"""

        # 如果没有问题
        if not result.issues:
            md += """## ✅ 未发现问题

恭喜！此文件未发现任何代码规范问题。

"""
            return md

        # 按严重程度分组
        errors = [i for i in result.issues if i.severity == Severity.ERROR]
        warnings = [i for i in result.issues if i.severity == Severity.WARNING]
        infos = [i for i in result.issues if i.severity == Severity.INFO]

        # 显示错误
        if errors:
            md += f"## ❌ 错误 ({len(errors)})\n\n"
            md += "以下问题必须修复，可能导致系统崩溃、安全漏洞或数据丢失：\n\n"
            for idx, issue in enumerate(errors, 1):
                md += self._format_issue_markdown(idx, issue)

        # 显示警告
        if warnings:
            md += f"## ⚠️ 警告 ({len(warnings)})\n\n"
            md += "以下问题强烈建议修复，影响代码质量、性能或可维护性：\n\n"
            for idx, issue in enumerate(warnings, 1):
                md += self._format_issue_markdown(idx, issue)

        # 显示提示
        if infos:
            md += f"## ℹ️ 提示 ({len(infos)})\n\n"
            md += "以下是代码改进建议：\n\n"
            for idx, issue in enumerate(infos, 1):
                md += self._format_issue_markdown(idx, issue)

        return md

    def _format_issue_markdown(self, index: int, issue: Issue) -> str:
        """
        格式化单个问题为 Markdown

        Args:
            index: 问题序号
            issue: 问题对象

        Returns:
            Markdown 格式的问题描述
        """
        md = f"### 问题 {index}\n\n"
        md += f"- **位置**: 第 {issue.line_start}"
        if issue.line_end != issue.line_start:
            # 计算实际行数（包含性：line_end - line_start + 1）
            line_count = issue.line_end - issue.line_start + 1
            md += f"-{issue.line_end} 行（共 {line_count} 行）\n"
        else:
            md += " 行\n"
        md += f"- **规则**: `{issue.rule_id}`\n"
        md += f"- **描述**: {issue.description}\n"
        md += f"- **建议**: {issue.suggestion}\n"

        # 如果有代码片段
        if issue.code_snippet:
            md += f"\n**问题代码**:\n```\n{issue.code_snippet}\n```\n"

        md += "\n---\n\n"
        return md

    def _get_git_report_title(self, git_info: GitInfo) -> str:
        """
        根据 Git 类型生成报告标题（Phase 4）

        Args:
            git_info: Git 检查信息

        Returns:
            报告标题字符串
        """
        if git_info.type == "staged":
            return "代码检查报告 - Git 暂存区"
        elif git_info.type == "unstaged":
            return "代码检查报告 - Git 工作区"
        elif git_info.type == "commit":
            return "代码检查报告 - Git Commit"
        elif git_info.type == "diff":
            return "代码检查报告 - Git Diff"
        else:
            return "代码检查报告"

    def _format_git_info_markdown(self, git_info: GitInfo) -> List[str]:
        """
        格式化 Git 信息为 Markdown（Phase 4）

        Args:
            git_info: Git 检查信息

        Returns:
            Markdown 格式的 Git 信息行列表
        """
        lines = []

        if git_info.type == "staged":
            lines.append("**检查类型**: Git 暂存区文件")
            if git_info.branch:
                lines.append(f"**当前分支**: {git_info.branch}")
            lines.append(f"**文件数量**: {git_info.files_changed} 个")

        elif git_info.type == "unstaged":
            lines.append("**检查类型**: Git 工作区修改文件")
            if git_info.branch:
                lines.append(f"**当前分支**: {git_info.branch}")
            lines.append(f"**文件数量**: {git_info.files_changed} 个")

        elif git_info.type == "commit":
            lines.append("**检查类型**: Git Commit 检查")
            if git_info.short_hash and git_info.message:
                # 截断过长的 commit message（只显示第一行）
                message_first_line = git_info.message.splitlines()[0] if git_info.message else ""
                if len(message_first_line) > 80:
                    message_first_line = message_first_line[:77] + "..."
                lines.append(f"**Commit**: `{git_info.short_hash}` - {message_first_line}")
            if git_info.author:
                lines.append(f"**作者**: {git_info.author}")
            if git_info.date:
                lines.append(f"**日期**: {git_info.date}")
            lines.append(f"**变更文件**: {git_info.files_changed} 个")

        elif git_info.type == "diff":
            lines.append("**检查类型**: Git Diff 检查")
            if git_info.commit1 and git_info.commit2:
                lines.append(f"**对比范围**: `{git_info.commit1}`...`{git_info.commit2}`")
            lines.append(f"**差异文件**: {git_info.files_changed} 个")

        return lines

    def _format_summary_markdown(self, batch_result: BatchCheckResult) -> str:
        """
        格式化批量检查结果为 Markdown

        Args:
            batch_result: 批量检查结果

        Returns:
            Markdown 格式的汇总报告内容
        """
        # 计算耗时
        duration = batch_result.get_duration_seconds()
        duration_str = f"{duration:.2f} 秒"
        if duration >= 60:
            duration_str = f"{duration / 60:.2f} 分钟"

        # Phase 4: 根据是否有 Git 信息决定标题
        if batch_result.git_info:
            title = self._get_git_report_title(batch_result.git_info)
        else:
            title = "代码检查汇总报告"

        # 构建 Markdown 内容
        md = f"# 📊 {title}\n\n"

        # Phase 4: 如果有 Git 信息，先显示 Git 信息
        if batch_result.git_info:
            git_lines = self._format_git_info_markdown(batch_result.git_info)
            for line in git_lines:
                md += f"{line}\n"
            md += f"**检查时间**: {batch_result.end_time}\n"
            md += "\n---\n\n"
        else:
            # 非 Git 检查，显示原有的检查 ID 和时间信息
            md += f"**检查 ID**: `{batch_result.check_id}`\n"
            md += f"**开始时间**: {batch_result.start_time}\n"
            md += f"**结束时间**: {batch_result.end_time}\n"
            md += f"**总耗时**: {duration_str}\n\n"

        md += "## 📈 检查概览\n\n"
        md += "| 统计项 | 数量 |\n"
        md += "|--------|------|\n"
        md += f"| 总文件数 | {batch_result.total_files} |\n"
        md += f"| 已检查文件 | {batch_result.checked_files} |\n"
        md += f"| 完成率 | {batch_result.get_completion_rate():.1f}% |\n"
        md += f"| **总问题数** | **{batch_result.total_issues}** |\n\n"

        # Phase 5: 显示审核模式统计
        diff_only_files = [r for r in batch_result.file_results if r.audit_mode == "diff-only"]
        full_files = [r for r in batch_result.file_results if r.audit_mode == "full"]

        if diff_only_files:
            md += "## 🎯 审核模式统计\n\n"
            md += "| 审核模式 | 文件数 | 占比 |\n"
            md += "|---------|-------|------|\n"
            md += f"| 🎯 Diff-Only（差异审核） | {len(diff_only_files)} | {len(diff_only_files) / batch_result.total_files * 100:.1f}% |\n"
            md += f"| 📄 全文件审核 | {len(full_files)} | {len(full_files) / batch_result.total_files * 100:.1f}% |\n\n"

            # 计算 diff-only 模式的整体效率提升
            total_audited = sum(r.audit_stats.get("audited_lines", 0) for r in diff_only_files if r.audit_stats)
            total_lines = sum(r.audit_stats.get("total_lines", 0) for r in diff_only_files if r.audit_stats)

            if total_lines > 0:
                overall_coverage = int(total_audited / total_lines * 100)
                efficiency_gain = 100 - overall_coverage

                md += f"**Diff-Only 模式效率**:\n"
                md += f"- 审核了 {total_audited:,}/{total_lines:,} 行代码\n"
                md += f"- 覆盖率: {overall_coverage}%\n"
                md += f"- 效率提升: 约 {efficiency_gain}% Token 节省\n\n"
                md += "---\n\n"


        md += "## 🔍 问题分布\n\n"
        md += "| 严重程度 | 数量 | 占比 |\n"
        md += "|---------|------|------|\n"
        md += f"| ❌ 错误 (ERROR) | {batch_result.total_errors} | {batch_result.total_errors / max(batch_result.total_issues, 1) * 100:.1f}% |\n"
        md += f"| ⚠️ 警告 (WARNING) | {batch_result.total_warnings} | {batch_result.total_warnings / max(batch_result.total_issues, 1) * 100:.1f}% |\n"
        md += f"| ℹ️ 提示 (INFO) | {batch_result.total_infos} | {batch_result.total_infos / max(batch_result.total_issues, 1) * 100:.1f}% |\n\n"
        md += "---\n\n"
        md += "## 📋 文件检查详情\n\n"

        # 按问题数量排序文件
        sorted_results = sorted(
            batch_result.file_results,
            key=lambda r: r.get_total_issues(),
            reverse=True
        )

        # 创建文件汇总表格
        md += "| 文件路径 | 状态 | 错误 | 警告 | 提示 | 总计 |\n"
        md += "|---------|------|------|------|------|------|\n"

        for result in sorted_results:
            status_icon = {
                "success": "✅",
                "failed": "❌",
                "skipped": "⏭️"
            }.get(result.status, "❓")

            # 截断过长的路径
            file_path = result.file_path
            if len(file_path) > 50:
                file_path = "..." + file_path[-47:]

            md += f"| `{file_path}` | {status_icon} | "
            md += f"{result.error_count} | {result.warning_count} | "
            md += f"{result.info_count} | **{result.get_total_issues()}** |\n"

        md += "\n---\n\n"

        # 显示有问题的文件详情
        files_with_issues = [r for r in sorted_results if r.get_total_issues() > 0]

        if not files_with_issues:
            md += """## ✅ 检查完成

所有文件均未发现问题，代码质量良好！

"""
            return md

        md += f"## 🔴 问题详情 (共 {len(files_with_issues)} 个文件有问题)\n\n"

        for file_result in files_with_issues:
            md += f"### 📄 {file_result.file_path}\n\n"
            md += f"**问题数**: {file_result.get_total_issues()} 个 "
            md += f"(❌ {file_result.error_count} ⚠️ {file_result.warning_count} "
            md += f"ℹ️ {file_result.info_count})\n\n"

            # 按严重程度分组问题
            errors = [i for i in file_result.issues if i.severity == Severity.ERROR]
            warnings = [i for i in file_result.issues if i.severity == Severity.WARNING]
            infos = [i for i in file_result.issues if i.severity == Severity.INFO]

            # 显示问题列表（简化版）
            for issue in errors[:3]:  # 最多显示 3 个错误
                md += f"- ❌ **第 {issue.line_start} 行**: {issue.description}\n"

            if len(errors) > 3:
                md += f"  _... 还有 {len(errors) - 3} 个错误_\n"

            for issue in warnings[:2]:  # 最多显示 2 个警告
                md += f"- ⚠️ **第 {issue.line_start} 行**: {issue.description}\n"

            if len(warnings) > 2:
                md += f"  _... 还有 {len(warnings) - 2} 个警告_\n"

            for issue in infos[:1]:  # 最多显示 1 个提示
                md += f"- ℹ️ **第 {issue.line_start} 行**: {issue.description}\n"

            if len(infos) > 1:
                md += f"  _... 还有 {len(infos) - 1} 个提示_\n"

            md += "\n"

        # 添加总结
        md += "---\n\n"
        md += "## 📝 建议\n\n"

        if batch_result.total_errors > 0:
            md += f"- ❌ 发现 **{batch_result.total_errors}** 个错误，请优先修复\n"

        if batch_result.total_warnings > 0:
            md += f"- ⚠️ 发现 **{batch_result.total_warnings}** 个警告，建议修复以提高代码质量\n"

        if batch_result.total_infos > 0:
            md += f"- ℹ️ 发现 **{batch_result.total_infos}** 个改进建议，可考虑优化\n"

        md += "\n## 📋 日志文件\n\n"
        md += "本次检查的详细执行日志已保存在：\n\n"
        md += f"- **日志文件**: `check.log`\n\n"
        md += "**日志文件用途**：\n"
        md += "- 记录完整的检查执行过程（文件扫描、规则加载、LLM 调用等）\n"
        md += "- 包含详细的 DEBUG 级别信息，便于问题排查\n"
        md += "- 记录所有警告和错误信息\n"
        md += "- 适用场景：\n"
        md += "  - 检查过程异常中断，需要了解中断原因\n"
        md += "  - LLM 调用失败或超时，需要查看详细错误信息\n"
        md += "  - 需要了解检查过程的性能数据（各文件耗时等）\n\n"

        md += "## 📁 报告文件组织\n\n"
        md += "为便于快速查看，报告文件已按问题分类存储：\n\n"

        # 统计有问题和无问题的文件数量
        files_with_issues = len([r for r in batch_result.file_results if r.get_total_issues() > 0])
        files_no_issues = len([r for r in batch_result.file_results if r.get_total_issues() == 0])

        md += f"- **有问题的文件** ({files_with_issues} 个): `files/with_issues/` 目录\n"
        md += f"- **无问题的文件** ({files_no_issues} 个): `files/no_issues/` 目录\n"
        md += f"- **日志文件**: `check.log`（详细执行过程）\n"
        md += "\n💡 **提示**: 优先查看 `files/with_issues/` 目录中的报告进行修复。如需排查问题，可查看 `check.log` 日志文件。\n"

        return md

    def _safe_path(self, file_path: str) -> str:
        """
        将文件路径转换为安全的短文件名

        使用"文件名_扩展名_哈希"格式，避免路径过长。
        哈希值基于完整路径计算，确保不同路径的同名文件不会冲突。

        Args:
            file_path: 原始文件路径

        Returns:
            安全的短文件名（格式：filename_ext_hash6）

        Examples:
            >>> gen = ReportGenerator()
            >>> gen._safe_path("autocoder/checker/core.py")
            'core_py_a1b2c3'
            >>> gen._safe_path("autocoder/plugins/core.py")
            'core_py_d4e5f6'  # 不同路径的同名文件，哈希不同
            >>> gen._safe_path("docs/二次开发记录.md")
            '二次开发记录_md_f7g8h9'
        """
        # 规范化路径分隔符（统一使用正斜杠），确保跨平台哈希一致性
        normalized_path = file_path.replace('\\', '/')

        # 提取文件名（不含路径）
        filename = os.path.basename(normalized_path)

        # 分离文件名和扩展名
        name_parts = filename.rsplit('.', 1)
        if len(name_parts) == 2:
            base_name, extension = name_parts
        else:
            base_name = name_parts[0]
            extension = ''

        # 清理文件名中的特殊字符（保留中文、字母、数字、下划线、短横线）
        def clean_name(s):
            # Windows 非法字符: < > : " / \ | ? *
            illegal_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
            result = s
            for char in illegal_chars:
                result = result.replace(char, '_')
            # 去除前后空格
            result = result.strip()
            # 如果为空，使用默认名称
            if not result:
                result = 'unnamed'
            return result

        clean_base_name = clean_name(base_name)
        clean_extension = clean_name(extension) if extension else ''

        # 计算完整路径的哈希值（6位十六进制）
        # 使用 MD5 算法，确保跨平台一致性
        path_hash = hashlib.md5(normalized_path.encode('utf-8')).hexdigest()[:6]

        # 组合文件名：文件名_扩展名_哈希
        if clean_extension:
            safe_filename = f"{clean_base_name}_{clean_extension}_{path_hash}"
        else:
            safe_filename = f"{clean_base_name}_{path_hash}"

        # 限制总长度（避免超长文件名）
        # Windows 文件名限制 255 字符，留余量给报告扩展名 (.json/.md)
        if len(safe_filename) > 100:
            # 如果太长，截断文件名部分（保留扩展名和哈希）
            max_base_len = 100 - len(clean_extension) - len(path_hash) - 2  # 2个下划线
            if clean_extension:
                safe_filename = f"{clean_base_name[:max_base_len]}_{clean_extension}_{path_hash}"
            else:
                safe_filename = f"{clean_base_name[:max_base_len]}_{path_hash}"

        return safe_filename

    # ==================== Excel 报告生成方法 ====================

    def _auto_adjust_column_width(self, ws, min_width: int = 8, max_width: int = 60) -> None:
        """
        自动调整工作表列宽

        Args:
            ws: openpyxl 工作表对象
            min_width: 最小列宽
            max_width: 最大列宽
        """
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        # 计算字符长度（中文字符按2计算，以适应Excel显示）
                        cell_value = str(cell.value)
                        # 考虑换行符，取最长行
                        if '\n' in cell_value:
                            cell_length = max(
                                sum(2 if ord(c) > 127 else 1 for c in line)
                                for line in cell_value.split('\n')
                            )
                        else:
                            cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass

            # 应用宽度限制
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width

    def _apply_header_style(self, ws, row: int = 1) -> None:
        """
        为指定行应用表头样式

        Args:
            ws: openpyxl 工作表对象
            row: 表头行号（从1开始）
        """
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

    def _generate_excel_file_report(self, result: FileCheckResult, output_path: str) -> None:
        """
        生成单文件 Excel 报告

        Args:
            result: 文件检查结果
            output_path: 输出文件路径

        Raises:
            RuntimeError: 如果报告生成失败
        """
        try:
            wb = Workbook()

            # ===== 概览工作表 =====
            ws_overview = wb.active
            ws_overview.title = "概览"

            # 概览数据
            overview_data = [
                ("文件路径", result.file_path),
                ("检查时间", result.check_time),
                ("检查状态", result.status),
            ]

            # 审核模式信息
            if result.audit_mode:
                mode_text = "Diff-Only（差异审核）" if result.audit_mode == "diff-only" else "全文件审核"
                if result.audit_stats and result.audit_mode == "diff-only":
                    audit_summary = result.get_audit_summary()
                    if audit_summary:
                        mode_text += f" - {audit_summary}"
                overview_data.append(("审核模式", mode_text))

            overview_data.extend([
                ("", ""),  # 空行分隔
                ("总问题数", result.get_total_issues()),
                ("错误 (ERROR)", result.error_count),
                ("警告 (WARNING)", result.warning_count),
                ("提示 (INFO)", result.info_count),
            ])

            if result.error_message:
                overview_data.append(("", ""))
                overview_data.append(("错误信息", result.error_message))

            # 写入概览数据
            for row_idx, (label, value) in enumerate(overview_data, start=1):
                ws_overview.cell(row=row_idx, column=1, value=label)
                ws_overview.cell(row=row_idx, column=2, value=value)
                # 标签列加粗
                if label:
                    ws_overview.cell(row=row_idx, column=1).font = Font(bold=True)
                # 状态列应用颜色
                if label == "检查状态":
                    status_fill = self.STATUS_FILLS.get(result.status)
                    if status_fill:
                        ws_overview.cell(row=row_idx, column=2).fill = status_fill

            self._auto_adjust_column_width(ws_overview)

            # ===== 问题详情工作表 =====
            ws_issues = wb.create_sheet("问题详情")

            # 表头
            headers = ["序号", "严重程度", "规则ID", "行号", "描述", "建议", "代码片段"]
            for col_idx, header in enumerate(headers, start=1):
                ws_issues.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(ws_issues, row=1)

            # 问题数据
            if result.issues:
                for row_idx, issue in enumerate(result.issues, start=2):
                    # 行号格式
                    if issue.line_start != issue.line_end:
                        line_range = f"{issue.line_start}-{issue.line_end}"
                    else:
                        line_range = str(issue.line_start)

                    ws_issues.cell(row=row_idx, column=1, value=row_idx - 1)

                    # 严重程度（带颜色）
                    severity_value = issue.severity.value.upper() if hasattr(issue.severity, 'value') else str(issue.severity).upper()
                    severity_cell = ws_issues.cell(row=row_idx, column=2, value=severity_value)
                    severity_key = issue.severity.value.lower() if hasattr(issue.severity, 'value') else str(issue.severity).lower()
                    severity_fill = self.SEVERITY_FILLS.get(severity_key)
                    if severity_fill:
                        severity_cell.fill = severity_fill

                    ws_issues.cell(row=row_idx, column=3, value=issue.rule_id)
                    ws_issues.cell(row=row_idx, column=4, value=line_range)
                    ws_issues.cell(row=row_idx, column=5, value=issue.description)
                    ws_issues.cell(row=row_idx, column=6, value=issue.suggestion)

                    # 代码片段（自动换行）
                    code_cell = ws_issues.cell(row=row_idx, column=7, value=issue.code_snippet or "")
                    code_cell.alignment = Alignment(wrap_text=True, vertical="top")

                # 冻结表头
                ws_issues.freeze_panes = "A2"
                # 启用筛选
                ws_issues.auto_filter.ref = ws_issues.dimensions
            else:
                # 无问题时显示提示
                ws_issues.cell(row=2, column=1, value="未发现问题")
                ws_issues.merge_cells('A2:G2')
                ws_issues.cell(row=2, column=1).alignment = Alignment(horizontal="center")

            self._auto_adjust_column_width(ws_issues)

            # 确保目录存在
            parent_dir = os.path.dirname(output_path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)

            # 保存文件
            wb.save(output_path)

            # 验证文件是否创建成功
            if not os.path.exists(output_path):
                raise RuntimeError(f"Excel 文件未创建成功: {output_path}")

            file_size = os.path.getsize(output_path)
            logger.debug(f"Excel 文件报告已保存: {output_path} (大小: {file_size} 字节)")

        except Exception as e:
            error_msg = f"生成 Excel 文件报告失败 {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise RuntimeError(error_msg) from e

    def _generate_excel_summary_report(
        self,
        results: List[FileCheckResult],
        output_path: str,
        git_info: Optional[GitInfo] = None
    ) -> None:
        """
        生成汇总 Excel 报告

        Args:
            results: 文件检查结果列表
            output_path: 输出文件路径
            git_info: Git 信息（可选）

        Raises:
            RuntimeError: 如果报告生成失败
        """
        try:
            wb = Workbook()

            # 统计数据
            total_files = len(results)
            checked_files = len([r for r in results if r.status == "success"])
            total_issues = sum(len(r.issues) for r in results)
            total_errors = sum(r.error_count for r in results)
            total_warnings = sum(r.warning_count for r in results)
            total_infos = sum(r.info_count for r in results)

            # ===== 汇总概览工作表 =====
            ws_overview = wb.active
            ws_overview.title = "汇总概览"

            overview_data = []

            # Git 信息
            if git_info:
                if git_info.type == "staged":
                    overview_data.append(("检查类型", "Git 暂存区文件"))
                elif git_info.type == "unstaged":
                    overview_data.append(("检查类型", "Git 工作区修改文件"))
                elif git_info.type == "commit":
                    overview_data.append(("检查类型", "Git Commit 检查"))
                    if git_info.short_hash:
                        msg = git_info.message[:50] + "..." if git_info.message and len(git_info.message) > 50 else (git_info.message or "")
                        overview_data.append(("Commit", f"{git_info.short_hash} - {msg}"))
                elif git_info.type == "diff":
                    overview_data.append(("检查类型", "Git Diff 检查"))
                    if git_info.commit1 and git_info.commit2:
                        overview_data.append(("对比范围", f"{git_info.commit1}...{git_info.commit2}"))

                if git_info.branch:
                    overview_data.append(("当前分支", git_info.branch))

            overview_data.extend([
                ("", ""),  # 空行分隔
                ("总文件数", total_files),
                ("已检查文件", checked_files),
                ("完成率", f"{checked_files / total_files * 100:.1f}%" if total_files > 0 else "0%"),
                ("", ""),
                ("总问题数", total_issues),
                ("错误 (ERROR)", total_errors),
                ("警告 (WARNING)", total_warnings),
                ("提示 (INFO)", total_infos),
            ])

            for row_idx, (label, value) in enumerate(overview_data, start=1):
                ws_overview.cell(row=row_idx, column=1, value=label)
                ws_overview.cell(row=row_idx, column=2, value=value)
                if label:
                    ws_overview.cell(row=row_idx, column=1).font = Font(bold=True)

            self._auto_adjust_column_width(ws_overview)

            # ===== 文件列表工作表 =====
            ws_files = wb.create_sheet("文件列表")

            file_headers = ["序号", "文件路径", "状态", "错误", "警告", "提示", "总计", "审核模式"]
            for col_idx, header in enumerate(file_headers, start=1):
                ws_files.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(ws_files, row=1)

            # 按严重程度优先排序：先按 ERROR 数量，再按 WARNING，最后 INFO
            sorted_results = sorted(
                results,
                key=lambda r: (r.error_count, r.warning_count, r.info_count),
                reverse=True
            )

            for row_idx, result in enumerate(sorted_results, start=2):
                ws_files.cell(row=row_idx, column=1, value=row_idx - 1)
                ws_files.cell(row=row_idx, column=2, value=result.file_path)

                # 状态（带颜色）
                status_cell = ws_files.cell(row=row_idx, column=3, value=result.status)
                status_fill = self.STATUS_FILLS.get(result.status)
                if status_fill:
                    status_cell.fill = status_fill

                ws_files.cell(row=row_idx, column=4, value=result.error_count)
                ws_files.cell(row=row_idx, column=5, value=result.warning_count)
                ws_files.cell(row=row_idx, column=6, value=result.info_count)
                ws_files.cell(row=row_idx, column=7, value=result.get_total_issues())
                ws_files.cell(row=row_idx, column=8, value=result.audit_mode or "-")

            ws_files.freeze_panes = "A2"
            ws_files.auto_filter.ref = ws_files.dimensions
            self._auto_adjust_column_width(ws_files)

            # ===== 问题分布工作表 =====
            ws_dist = wb.create_sheet("问题分布")

            # 按规则ID统计
            rule_stats: Dict[str, Dict[str, int]] = {}
            for result in results:
                for issue in result.issues:
                    if issue.rule_id not in rule_stats:
                        rule_stats[issue.rule_id] = {"error": 0, "warning": 0, "info": 0}
                    severity_key = issue.severity.value.lower() if hasattr(issue.severity, 'value') else str(issue.severity).lower()
                    rule_stats[issue.rule_id][severity_key] += 1

            dist_headers = ["规则ID", "ERROR", "WARNING", "INFO", "总计"]
            for col_idx, header in enumerate(dist_headers, start=1):
                ws_dist.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(ws_dist, row=1)

            # 按总数排序规则
            sorted_rules = sorted(rule_stats.items(), key=lambda x: sum(x[1].values()), reverse=True)

            for row_idx, (rule_id, counts) in enumerate(sorted_rules, start=2):
                total = sum(counts.values())
                ws_dist.cell(row=row_idx, column=1, value=rule_id)
                ws_dist.cell(row=row_idx, column=2, value=counts["error"])
                ws_dist.cell(row=row_idx, column=3, value=counts["warning"])
                ws_dist.cell(row=row_idx, column=4, value=counts["info"])
                ws_dist.cell(row=row_idx, column=5, value=total)

            if rule_stats:
                ws_dist.freeze_panes = "A2"
            self._auto_adjust_column_width(ws_dist)

            # ===== 所有问题工作表 =====
            ws_all = wb.create_sheet("所有问题")

            all_headers = ["序号", "文件路径", "严重程度", "规则ID", "行号", "描述", "建议"]
            for col_idx, header in enumerate(all_headers, start=1):
                ws_all.cell(row=1, column=col_idx, value=header)
            self._apply_header_style(ws_all, row=1)

            issue_idx = 1
            for result in sorted_results:
                for issue in result.issues:
                    row_idx = issue_idx + 1

                    # 行号格式
                    if issue.line_start != issue.line_end:
                        line_range = f"{issue.line_start}-{issue.line_end}"
                    else:
                        line_range = str(issue.line_start)

                    ws_all.cell(row=row_idx, column=1, value=issue_idx)
                    ws_all.cell(row=row_idx, column=2, value=result.file_path)

                    # 严重程度（带颜色）
                    severity_value = issue.severity.value.upper() if hasattr(issue.severity, 'value') else str(issue.severity).upper()
                    severity_cell = ws_all.cell(row=row_idx, column=3, value=severity_value)
                    severity_key = issue.severity.value.lower() if hasattr(issue.severity, 'value') else str(issue.severity).lower()
                    severity_fill = self.SEVERITY_FILLS.get(severity_key)
                    if severity_fill:
                        severity_cell.fill = severity_fill

                    ws_all.cell(row=row_idx, column=4, value=issue.rule_id)
                    ws_all.cell(row=row_idx, column=5, value=line_range)

                    # 描述和建议（自动换行）
                    desc_cell = ws_all.cell(row=row_idx, column=6, value=issue.description)
                    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

                    suggest_cell = ws_all.cell(row=row_idx, column=7, value=issue.suggestion)
                    suggest_cell.alignment = Alignment(wrap_text=True, vertical="top")

                    issue_idx += 1

            if issue_idx > 1:
                ws_all.freeze_panes = "A2"
                ws_all.auto_filter.ref = ws_all.dimensions
            self._auto_adjust_column_width(ws_all)

            # 确保目录存在
            parent_dir = os.path.dirname(output_path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)

            # 保存文件
            wb.save(output_path)

            # 验证文件是否创建成功
            if not os.path.exists(output_path):
                raise RuntimeError(f"Excel 汇总报告未创建成功: {output_path}")

            file_size = os.path.getsize(output_path)
            logger.info(f"Excel 汇总报告已保存: {output_path} (大小: {file_size} 字节)")

        except Exception as e:
            error_msg = f"生成 Excel 汇总报告失败 {output_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise RuntimeError(error_msg) from e
